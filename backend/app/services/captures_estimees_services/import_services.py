"""
SIGPA — Module « Captures estimées »
Import du classeur « Données_estimées_<année> ».

Structure attendue (feuille unique) :
  • Deux grandes sections empilées verticalement, séparées par un bandeau :
        1. « Captures (kg) »   → tonnages estimés
        2. « Valeurs (f.cfa) » → valeurs estimées
  • Chaque section est découpée horizontalement en BLOCS de 15 colonnes,
    un bloc par engin de pêche. Le bloc de tête (sans titre d'engin, ou dont
    le titre n'est pas un libellé d'engin plausible) correspond au cumul
    TOUS ENGINS et est marqué `agrege`.
  • Dans un bloc :
        - en-tête : « Espèces | Janvier | … | Décembre | Total »
        - lignes espèces  : nom + 12 valeurs mensuelles + total
        - lignes de synthèse : « Efforts (jr) », « Nbre débarq. »,
          « Taux échant. », « CPUE (kg/jr) », « Captures (kg) »

Principes de robustesse (classeur d'analyste souvent irrégulier) :
  • La section (kg vs f.cfa) est déterminée par la POSITION du bandeau
    « Valeurs (f.cfa) » — signal fiable — et non par un balayage de colonne.
  • Seuls les blocs proches de leur bandeau de section sont importés ; les
    tableaux annexes situés plus bas (production par groupe, blocs tuilés
    non standard) sont IGNORÉS et signalés, jamais importés « au jugé ».
  • Un titre d'engin purement numérique est rejeté (→ bloc agrégé).
  • Les erreurs sont isolées ligne par ligne (savepoints) sans interrompre
    l'import ; chaque anomalie est renvoyée dans `erreurs`.

Dépendance : openpyxl.
"""

from __future__ import annotations

import re
from io import BytesIO
from typing import Optional

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import captures_estimees as models
from app.schemas import captures_estimees as schemas
from app.services.captures_estimees_services import main_services as service

MOIS = schemas.MOIS_LIBELLES
_MOIS_LOWER = {m.lower() for m in MOIS}

LIBELLES_SYNTHESE = {
    "efforts (jr)",
    "captures (kg)",
    "cpue (kg/jr)",
    "nbre débarq.",
    "nbre debarq.",
    "taux échant.",
    "taux echant.",
    "captures (tonnes)",
    "valeurs (f.cfa)",
    "valeurs (,000 f.cfa)",
    "total (,000. f.cfa)",
}
TITRES_SECTION = {"captures (kg)", "valeurs (f.cfa)"}

# Fenêtre (en lignes) autour d'un bandeau de section dans laquelle un en-tête
# « Espèces » est considéré comme appartenant à cette section.
FENETRE_SECTION = 8

# Classification biologique par défaut (complète le référentiel Espèces).
GROUPES = {
    models.GroupeEspece.PELAGIQUE: {
        "Banane de Mer",
        "Barbillon",
        "Bécune",
        "Carangue",
        "Carpe",
        "Ceinture",
        "Maquéreau",
        "Mulet",
        "Raie",
        "Requin",
        "Sardine",
        "Thon",
        "Turbo",
    },
    models.GroupeEspece.DEMERSAL: {
        "Bars",
        "Bossu",
        "Capitaine",
        "Disque",
        "Divers",
        "Dorade Grise",
        "Dorade Rose",
        "Machoiron de mer",
        "Merou",
        "Rouge",
        "Sole",
    },
    models.GroupeEspece.CRUSTACE: {"Crabe", "Crevette", "Langouste"},
}
_ESPECE_GROUPE = {nom: g for g, noms in GROUPES.items() for nom in noms}


# ---------------------------------------------------------------------------
# Helpers de lecture
# ---------------------------------------------------------------------------
def _norm(v) -> str:
    return str(v).strip().lower() if v is not None else ""


def _num(v) -> Optional[float]:
    """Convertit une cellule en float ; None si vide/non numérique."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        f = float(v)
        return None if f != f else f  # écarte NaN
    try:
        return float(str(v).replace(" ", "").replace(",", "."))
    except ValueError:
        return None


def _annee_depuis_nom(nom_fichier: str, defaut: int) -> int:
    m = re.search(r"(20\d{2})", nom_fichier or "")
    return int(m.group(1)) if m else defaut


def _est_titre_engin(v) -> bool:
    n = _norm(v)
    if not n or n in _MOIS_LOWER or n in TITRES_SECTION or n in ("kg", "f.cfa"):
        return False
    if _num(v) is not None:  # purement numérique → pas un engin
        return False
    return any(ch.isalpha() for ch in n)


# ---------------------------------------------------------------------------
# Repérage des bandeaux de section
# ---------------------------------------------------------------------------
def _localiser_reperes(ws) -> tuple[int, int]:
    """
    Renvoie (row_fcfa, row_prod) :
      row_fcfa = ligne du bandeau « Valeurs (f.cfa) » (kg au-dessus, f.cfa en-dessous)
      row_prod = ligne « Production par Groupe… » (borne basse d'import)
    """
    row_fcfa = ws.max_row + 1
    row_prod = ws.max_row + 1
    for r in range(1, ws.max_row + 1):
        labels = [_norm(ws.cell(r, c).value) for c in range(1, 6)]
        if row_fcfa > ws.max_row and any(l == "valeurs (f.cfa)" for l in labels):
            row_fcfa = r
        if any(l.startswith("production par groupe") for l in labels):
            row_prod = r
            break
    return row_fcfa, row_prod


def _detecter_blocs(ws, row_fcfa: int, row_prod: int) -> tuple[list[dict], list[dict]]:
    """
    Retourne (blocs_retenus, blocs_ignores). Un bloc est retenu s'il est proche
    de son bandeau de section ; sinon il est signalé mais non importé.
    """
    retenus, ignores = [], []
    for r in range(1, row_prod):
        for c in range(1, ws.max_column + 1):
            if _norm(ws.cell(r, c).value) != "espèces":
                continue
            section = "fcfa" if r > row_fcfa else "kg"

            # proximité du bandeau de section
            proche = (
                (r < row_fcfa) if section == "kg" else (r <= row_fcfa + FENETRE_SECTION)
            )

            engin = ""
            for rr in range(r - 1, max(r - 4, 0), -1):
                for cc in range(c, c + 14):
                    if _est_titre_engin(ws.cell(rr, cc).value):
                        engin = str(ws.cell(rr, cc).value).strip()
                        break
                if engin:
                    break

            bloc = dict(
                row_entete=r, col=c, engin=engin, section=section, agrege=(engin == "")
            )
            (retenus if proche else ignores).append(bloc)
    return retenus, ignores


# ---------------------------------------------------------------------------
# Point d'entrée
# ---------------------------------------------------------------------------
def importer_donnees_estimees(
    db: Session, contenu: bytes, nom_fichier: str, annee_defaut: int
) -> schemas.ImportResultat:
    wb = load_workbook(BytesIO(contenu), data_only=True)
    ws = wb.active
    annee = _annee_depuis_nom(nom_fichier, annee_defaut)

    res = schemas.ImportResultat(
        lignes_lues=0,
        captures_importees=0,
        efforts_importes=0,
        engins_crees=0,
        especes_creees=0,
        erreurs=[],
    )

    row_fcfa, row_prod = _localiser_reperes(ws)
    retenus, ignores = _detecter_blocs(ws, row_fcfa, row_prod)

    for b in ignores:  # transparence : on dit ce qu'on n'importe pas
        res.erreurs.append(
            schemas.LigneErreur(
                feuille=ws.title,
                reference=f"Bloc ligne {b['row_entete']} col {b['col']}",
                message=(
                    "Tableau hors des deux sections canoniques "
                    "(kg / f.cfa) — ignoré, à vérifier manuellement."
                ),
            )
        )

    agrege_vus: set[str] = (
        set()
    )  # un seul bloc « TOTAL » (le plus à gauche) par section
    for bloc in retenus:
        if bloc["agrege"]:
            if bloc["section"] in agrege_vus:
                res.erreurs.append(
                    schemas.LigneErreur(
                        feuille=ws.title,
                        reference=f"Bloc agrégé ligne {bloc['row_entete']} col {bloc['col']}",
                        message="Second bloc agrégé de la section — ignoré (évite le double comptage).",
                    )
                )
                continue
            agrege_vus.add(bloc["section"])
        _importer_bloc(db, ws, bloc, annee, nom_fichier, row_prod, res)

    db.commit()
    # Succès « partiel » accepté : l'import réussit dès qu'au moins une capture
    # est enregistrée ; les anomalies restent listées dans `erreurs` pour revue.
    res.succes = res.captures_importees > 0
    return res


def _importer_bloc(db, ws, bloc, annee, nom_fichier, row_prod, res) -> None:
    col = bloc["col"]
    engin_libelle = bloc["engin"] or "TOTAL (tous engins)"
    try:
        engin, cree = service.get_ou_cree_engin(
            db, engin_libelle, agrege=bloc["agrege"]
        )
        if cree:
            res.engins_crees += 1
    except Exception as exc:
        res.erreurs.append(
            schemas.LigneErreur(
                feuille=ws.title,
                reference=f"Bloc col.{col}",
                message=f"Engin illisible : {exc}",
            )
        )
        return

    row = bloc["row_entete"] + 1
    blancs = 0
    while row < row_prod:
        libelle = ws.cell(row, col).value
        n = _norm(libelle)

        ligne_vide = n == "" and all(
            _num(ws.cell(row, col + 1 + k).value) is None for k in range(12)
        )
        if ligne_vide:
            blancs += 1
            if blancs >= 2:
                break
            row += 1
            continue
        blancs = 0

        if n == "espèces":  # entête d'un autre bloc → fin du bloc courant
            break
        if n in LIBELLES_SYNTHESE:
            _lire_ligne_synthese(db, ws, row, col, annee, engin, res)
            row += 1
            continue
        if n == "":  # libellé d'espèce absent (ligne de total isolée) → ignorée
            row += 1
            continue

        res.lignes_lues += 1
        reference = f"{engin_libelle} / {libelle} / {bloc['section']}"
        sp = db.begin_nested()  # savepoint : isole l'erreur de la ligne
        try:
            espece, cree = service.get_ou_cree_espece(
                db, str(libelle).strip(), _ESPECE_GROUPE.get(str(libelle).strip())
            )
            if cree:
                res.especes_creees += 1

            for m in range(12):
                valeur = _num(ws.cell(row, col + 1 + m).value) or 0.0
                if bloc["section"] == "kg":
                    service.upsert_capture(
                        db,
                        annee,
                        m + 1,
                        engin.id,
                        espece.id,
                        capture_kg=valeur,
                        valeur_fcfa=None,
                        source=f"Import {nom_fichier}",
                    )
                else:  # f.cfa : complète (ou crée) la cellule
                    c = _get_cellule(db, annee, m + 1, engin.id, espece.id)
                    if c is None:
                        c = models.CaptureEstimee(
                            annee=annee,
                            mois=m + 1,
                            engin_id=engin.id,
                            espece_id=espece.id,
                            capture_kg=0.0,
                            source=f"Import {nom_fichier}",
                        )
                        db.add(c)
                        db.flush()
                    c.valeur_fcfa = valeur
            res.captures_importees += 1
            sp.commit()
        except Exception as exc:
            sp.rollback()
            res.erreurs.append(
                schemas.LigneErreur(
                    feuille=ws.title, reference=reference, message=str(exc)
                )
            )
        row += 1


def _lire_ligne_synthese(db, ws, row, col, annee, engin, res) -> None:
    """Alimente les efforts (jours, débarquements, taux d'échant.) par mois."""
    n = _norm(ws.cell(row, col).value)
    champ = {
        "efforts (jr)": "efforts_jours",
        "nbre débarq.": "nombre_debarquements",
        "nbre debarq.": "nombre_debarquements",
        "taux échant.": "taux_echantillonnage",
        "taux echant.": "taux_echantillonnage",
    }.get(n)
    if not champ:
        return
    touche = False
    for m in range(12):
        val = _num(ws.cell(row, col + 1 + m).value)
        if val is None:
            continue
        e = _get_ou_cree_effort(db, annee, m + 1, engin.id)
        if champ == "nombre_debarquements":
            e.nombre_debarquements = int(round(val))
        else:
            setattr(e, champ, val)
        touche = True
    if touche:
        res.efforts_importes += 1


# --- accès directs (sans commit prématuré) ----------------------------------
def _get_cellule(db, annee, mois, engin_id, espece_id):
    return db.execute(
        select(models.CaptureEstimee).where(
            models.CaptureEstimee.annee == annee,
            models.CaptureEstimee.mois == mois,
            models.CaptureEstimee.engin_id == engin_id,
            models.CaptureEstimee.espece_id == espece_id,
        )
    ).scalar_one_or_none()


def _get_ou_cree_effort(db, annee, mois, engin_id):
    e = db.execute(
        select(models.EffortEstime).where(
            models.EffortEstime.annee == annee,
            models.EffortEstime.mois == mois,
            models.EffortEstime.engin_id == engin_id,
        )
    ).scalar_one_or_none()
    if e is None:
        e = models.EffortEstime(annee=annee, mois=mois, engin_id=engin_id)
        db.add(e)
        db.flush()
    return e
