"""
SIGPA — Module « Captures estimées »
Export multi-format : Excel (openpyxl), CSV, JSON, PDF (reportlab).
Retourne (contenu_bytes, media_type, nom_fichier).
"""

from __future__ import annotations

import csv
import io
import json
from datetime import datetime

from sqlalchemy.orm import Session

from app.schemas import captures_estimees as schemas
from app.services.captures_estimees_services import main_services as service

_ENTETES = [
    "Année",
    "Mois",
    "Engin",
    "Espèce",
    "Groupe",
    "Capture (kg)",
    "Capture (t)",
    "Valeur (f.CFA)",
    "Source",
]


def _lignes(db: Session, filtre: schemas.CaptureFiltre) -> list[schemas.CaptureOut]:
    f = filtre.model_copy(update={"page": 1, "taille_page": 500})
    out: list[schemas.CaptureOut] = []
    while True:
        page = service.lister_captures(db, f)
        out.extend(page.elements)
        if f.page * f.taille_page >= page.total:
            break
        f = f.model_copy(update={"page": f.page + 1})
    return out


def exporter(db: Session, req: schemas.ExportRequete) -> tuple[bytes, str, str]:
    donnees = _lignes(db, req.filtre)
    horodatage = datetime.now().strftime("%Y%m%d_%H%M")
    base = f"captures_estimees_{horodatage}"

    if req.format == schemas.FormatExport.json:
        payload = [d.model_dump() for d in donnees]
        return (
            json.dumps(payload, ensure_ascii=False, indent=2, default=str).encode(
                "utf-8"
            ),
            "application/json",
            f"{base}.json",
        )

    if req.format == schemas.FormatExport.csv:
        buf = io.StringIO()
        w = csv.writer(buf, delimiter=";")
        w.writerow(_ENTETES)
        for d in donnees:
            w.writerow(
                [
                    d.annee,
                    d.mois_libelle,
                    d.engin_libelle,
                    d.espece_nom,
                    d.espece_groupe or "",
                    f"{d.capture_kg:.3f}",
                    f"{d.capture_tonnes:.3f}",
                    f"{d.valeur_fcfa:.0f}",
                    d.source or "",
                ]
            )
        return buf.getvalue().encode("utf-8-sig"), "text/csv", f"{base}.csv"

    if req.format == schemas.FormatExport.excel:
        return (
            _excel(donnees),
            ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            f"{base}.xlsx",
        )

    return _pdf(donnees, req.filtre), "application/pdf", f"{base}.pdf"


def _excel(donnees: list[schemas.CaptureOut]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Captures estimées"
    entete_fill = PatternFill("solid", fgColor="1565C0")
    entete_font = Font(color="FFFFFF", bold=True)
    for j, titre in enumerate(_ENTETES, start=1):
        c = ws.cell(1, j, titre)
        c.fill = entete_fill
        c.font = entete_font
        c.alignment = Alignment(horizontal="center")
    for i, d in enumerate(donnees, start=2):
        ws.cell(i, 1, d.annee)
        ws.cell(i, 2, d.mois_libelle)
        ws.cell(i, 3, d.engin_libelle)
        ws.cell(i, 4, d.espece_nom)
        ws.cell(i, 5, d.espece_groupe or "")
        ws.cell(i, 6, round(d.capture_kg, 3))
        ws.cell(i, 7, round(d.capture_tonnes, 3))
        ws.cell(i, 8, round(d.valeur_fcfa, 0))
        ws.cell(i, 9, d.source or "")
    for col, largeur in zip("ABCDEFGHI", (8, 12, 24, 18, 12, 14, 12, 18, 22)):
        ws.column_dimensions[col].width = largeur
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _pdf(donnees: list[schemas.CaptureOut], filtre: schemas.CaptureFiltre) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), title="Captures estimées")
    styles = getSampleStyleSheet()
    elems = [
        Paragraph("SIGPA — Captures estimées", styles["Title"]),
        Paragraph(
            f"Année : {filtre.annee or 'toutes'} — "
            f"{len(donnees)} enregistrement(s) — "
            f"édité le {datetime.now():%d/%m/%Y %H:%M}",
            styles["Normal"],
        ),
        Spacer(1, 12),
    ]
    lignes = [_ENTETES[:8]]
    total_kg = total_val = 0.0
    for d in donnees[:2000]:  # garde-fou volume
        total_kg += d.capture_kg
        total_val += d.valeur_fcfa
        lignes.append(
            [
                d.annee,
                d.mois_libelle,
                d.engin_libelle,
                d.espece_nom,
                d.espece_groupe or "",
                f"{d.capture_kg:,.0f}",
                f"{d.capture_tonnes:,.2f}",
                f"{d.valeur_fcfa:,.0f}",
            ]
        )
    lignes.append(
        [
            "",
            "",
            "",
            "",
            "TOTAL",
            f"{total_kg:,.0f}",
            f"{total_kg/1000:,.2f}",
            f"{total_val:,.0f}",
        ]
    )
    t = Table(lignes, repeatRows=1)
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1565C0")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -2),
                    [colors.white, colors.HexColor("#EEF3FA")],
                ),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#DDE7F5")),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ("ALIGN", (5, 0), (-1, -1), "RIGHT"),
            ]
        )
    )
    elems.append(t)
    doc.build(elems)
    return buf.getvalue()
