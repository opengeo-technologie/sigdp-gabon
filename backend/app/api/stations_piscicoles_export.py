# -*- coding: utf-8 -*-
"""
Router Export — Module Stations Piscicoles (SIGDP-GABON)
Endpoints POST-only générant les fichiers côté serveur, sur TOUT le registre
filtré (pas de pagination) :

- POST /stations-piscicoles/export/excel   -> .xlsx (openpyxl)
- POST /stations-piscicoles/export/csv     -> .csv texte (BOM UTF-8, séparateur ';')
- POST /stations-piscicoles/export/json    -> .json
- POST /stations-piscicoles/export/pdf     -> .pdf (reportlab, registre officiel)

Dépendances : pip install openpyxl reportlab
"""

import csv
import io
import json
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import or_
from sqlalchemy.orm import Session

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
)

from app.database import get_db  # à ajuster selon l'arborescence
from app.models.stations_piscicole import StationPiscicole

router = APIRouter(
    prefix="/api/stations-piscicoles", tags=["Stations piscicoles - Export"]
)


# ---------------------------------------------------------------------------
# Libellés français (miroir du frontend)
# ---------------------------------------------------------------------------

TYPE_STATION_LABELS = {
    "ETANGS": "Étangs",
    "BACS_HORS_SOL": "Bacs hors-sol",
    "CAGES_FLOTTANTES": "Cages flottantes",
    "ECLOSERIE": "Écloserie",
    "MIXTE": "Mixte",
}
SOURCE_EAU_LABELS = {
    "FORAGE": "Forage",
    "RIVIERE": "Rivière",
    "LAC": "Lac",
    "RESEAU": "Réseau d'eau",
    "AUTRE": "Autre",
}
TYPE_PROMOTEUR_LABELS = {
    "PRIVE": "Privé",
    "COOPERATIVE": "Coopérative",
    "ETATIQUE": "Étatique",
    "PROJET": "Projet",
}
STATUT_LABELS = {
    "EN_CONSTRUCTION": "En construction",
    "ACTIVE": "Active",
    "SUSPENDUE": "Suspendue",
    "FERMEE": "Fermée",
}

ENTETES = [
    "Code",
    "Nom",
    "Province",
    "Département",
    "Localité",
    "Adresse",
    "Latitude",
    "Longitude",
    "Type de station",
    "Superficie (m²)",
    "Nombre de bassins",
    "Capacité (t/an)",
    "Source d'eau",
    "Espèces",
    "Promoteur",
    "Contact promoteur",
    "Type promoteur",
    "Statut",
    "Numéro agrément",
    "Date agrément",
    "Date expiration agrément",
    "Observations",
]


class ExportRequest(BaseModel):
    """Mêmes filtres que /list, sans pagination : l'export couvre tout."""

    search: Optional[str] = None
    province: Optional[str] = None
    type_station: Optional[str] = None
    statut: Optional[str] = None
    espece: Optional[str] = None


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _stations_filtrees(req: ExportRequest, db: Session) -> list[StationPiscicole]:
    query = db.query(StationPiscicole)
    if req.search:
        terme = f"%{req.search.strip()}%"
        query = query.filter(
            or_(
                StationPiscicole.nom.ilike(terme),
                StationPiscicole.code_station.ilike(terme),
                StationPiscicole.promoteur_nom.ilike(terme),
                StationPiscicole.localite.ilike(terme),
            )
        )
    if req.province:
        query = query.filter(StationPiscicole.province == req.province)
    if req.type_station:
        query = query.filter(StationPiscicole.type_station == req.type_station)
    if req.statut:
        query = query.filter(StationPiscicole.statut == req.statut)
    if req.espece:
        query = query.filter(StationPiscicole.especes_elevees.ilike(f"%{req.espece}%"))
    return query.order_by(StationPiscicole.province, StationPiscicole.nom).all()


def _date_fr(d) -> str:
    return d.strftime("%d/%m/%Y") if d else ""


def _station_vers_ligne(s: StationPiscicole) -> list:
    return [
        s.code_station,
        s.nom,
        s.province,
        s.departement or "",
        s.localite or "",
        s.adresse or "",
        s.latitude if s.latitude is not None else "",
        s.longitude if s.longitude is not None else "",
        TYPE_STATION_LABELS.get(s.type_station.name, s.type_station.name),
        s.superficie_totale if s.superficie_totale is not None else "",
        s.nombre_bassins if s.nombre_bassins is not None else "",
        s.capacite_production if s.capacite_production is not None else "",
        SOURCE_EAU_LABELS.get(s.source_eau.name, "") if s.source_eau else "",
        s.especes_elevees.replace(",", ", ") if s.especes_elevees else "",
        s.promoteur_nom,
        s.promoteur_contact or "",
        TYPE_PROMOTEUR_LABELS.get(s.promoteur_type.name, ""),
        STATUT_LABELS.get(s.statut.name, s.statut.name),
        s.numero_agrement or "",
        _date_fr(s.date_agrement),
        _date_fr(s.date_expiration_agrement),
        s.observations or "",
    ]


def _nom_fichier(extension: str) -> str:
    horodatage = datetime.utcnow().strftime("%Y%m%d")
    return f"stations_piscicoles_{horodatage}.{extension}"


def _reponse_fichier(tampon: io.BytesIO, media_type: str, extension: str):
    tampon.seek(0)
    return StreamingResponse(
        tampon,
        media_type=media_type,
        headers={
            "Content-Disposition": f'attachment; filename="{_nom_fichier(extension)}"'
        },
    )


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------


@router.post("/export/excel")
def exporter_excel(req: ExportRequest, db: Session = Depends(get_db)):
    stations = _stations_filtrees(req, db)

    wb = Workbook()
    ws = wb.active
    ws.title = "Stations piscicoles"

    entete_fill = PatternFill("solid", start_color="1B5E20")
    entete_font = Font(bold=True, color="FFFFFF")
    bordure = Border(*[Side(style="thin", color="BDBDBD")] * 4)

    for col, nom in enumerate(ENTETES, start=1):
        cellule = ws.cell(row=1, column=col, value=nom)
        cellule.fill = entete_fill
        cellule.font = entete_font
        cellule.alignment = Alignment(horizontal="center", vertical="center")
        cellule.border = bordure

    for rangee, station in enumerate(stations, start=2):
        for col, valeur in enumerate(_station_vers_ligne(station), start=1):
            cellule = ws.cell(row=rangee, column=col, value=valeur)
            cellule.border = bordure

    # Largeurs auto (plafonnées) + figer la ligne d'en-tête + filtre auto
    for col, nom in enumerate(ENTETES, start=1):
        lettre = ws.cell(row=1, column=col).column_letter
        largeur = (
            max(
                len(nom),
                *(
                    len(str(ws.cell(row=r, column=col).value or ""))
                    for r in range(2, min(len(stations) + 2, 200))
                ),
            )
            if stations
            else len(nom)
        )
        ws.column_dimensions[lettre].width = min(largeur + 3, 45)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    tampon = io.BytesIO()
    wb.save(tampon)
    return _reponse_fichier(
        tampon,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "xlsx",
    )


# ---------------------------------------------------------------------------
# CSV / texte
# ---------------------------------------------------------------------------


@router.post("/export/csv")
def exporter_csv(req: ExportRequest, db: Session = Depends(get_db)):
    stations = _stations_filtrees(req, db)

    tampon_texte = io.StringIO()
    scribe = csv.writer(
        tampon_texte, delimiter=";", quoting=csv.QUOTE_MINIMAL, lineterminator="\r\n"
    )
    scribe.writerow(ENTETES)
    for station in stations:
        scribe.writerow(_station_vers_ligne(station))

    # BOM UTF-8 pour ouverture directe dans Excel FR
    tampon = io.BytesIO(("\ufeff" + tampon_texte.getvalue()).encode("utf-8"))
    return _reponse_fichier(tampon, "text/csv; charset=utf-8", "csv")


# ---------------------------------------------------------------------------
# JSON
# ---------------------------------------------------------------------------


@router.post("/export/json")
def exporter_json(req: ExportRequest, db: Session = Depends(get_db)):
    stations = _stations_filtrees(req, db)

    donnees = {
        "module": "stations_piscicoles",
        "exporte_le": datetime.utcnow().isoformat() + "Z",
        "filtres": req.model_dump(exclude_none=True),
        "total": len(stations),
        "stations": [
            {
                "code_station": s.code_station,
                "nom": s.nom,
                "province": s.province,
                "departement": s.departement,
                "localite": s.localite,
                "adresse": s.adresse,
                "latitude": s.latitude,
                "longitude": s.longitude,
                "type_station": s.type_station.name,
                "superficie_totale": s.superficie_totale,
                "nombre_bassins": s.nombre_bassins,
                "capacite_production": s.capacite_production,
                "source_eau": s.source_eau.name if s.source_eau else None,
                "especes_elevees": s.especes_elevees,
                "promoteur_nom": s.promoteur_nom,
                "promoteur_contact": s.promoteur_contact,
                "promoteur_type": s.promoteur_type.name,
                "statut": s.statut.name,
                "numero_agrement": s.numero_agrement,
                "date_agrement": (
                    s.date_agrement.isoformat() if s.date_agrement else None
                ),
                "date_expiration_agrement": (
                    s.date_expiration_agrement.isoformat()
                    if s.date_expiration_agrement
                    else None
                ),
                "observations": s.observations,
            }
            for s in stations
        ],
    }

    tampon = io.BytesIO(
        json.dumps(donnees, ensure_ascii=False, indent=2).encode("utf-8")
    )
    return _reponse_fichier(tampon, "application/json; charset=utf-8", "json")


# ---------------------------------------------------------------------------
# PDF (reportlab) — registre officiel A4 paysage
# ---------------------------------------------------------------------------

COULEURS_STATUT_PDF = {
    "ACTIVE": colors.HexColor("#2e7d32"),
    "SUSPENDUE": colors.HexColor("#ef6c00"),
    "FERMEE": colors.HexColor("#c62828"),
    "EN_CONSTRUCTION": colors.HexColor("#546e7a"),
}


def _entete_pied_pdf(canvas, doc):
    """En-tête Ministère trois colonnes + pied de page numéroté."""
    canvas.saveState()
    largeur, hauteur = landscape(A4)

    # Colonne gauche
    canvas.setFont("Helvetica-Bold", 8)
    canvas.drawString(15 * mm, hauteur - 12 * mm, "RÉPUBLIQUE GABONAISE")
    canvas.setFont("Helvetica-Oblique", 7)
    canvas.drawString(15 * mm, hauteur - 16 * mm, "Union — Travail — Justice")

    # Colonne centrale
    canvas.setFont("Helvetica-Bold", 8)
    canvas.drawCentredString(
        largeur / 2, hauteur - 10 * mm, "MINISTÈRE DE LA MER, DE LA PÊCHE"
    )
    canvas.drawCentredString(largeur / 2, hauteur - 14 * mm, "ET DE L'ÉCONOMIE BLEUE")
    canvas.setFont("Helvetica", 7)
    canvas.drawCentredString(largeur / 2, hauteur - 18 * mm, "SIGDP-GABON")

    # Colonne droite
    canvas.setFont("Helvetica", 7)
    canvas.drawRightString(
        largeur - 15 * mm,
        hauteur - 12 * mm,
        f"Édité le {datetime.utcnow().strftime('%d/%m/%Y')}",
    )

    # Pied de page
    canvas.setFont("Helvetica", 7)
    canvas.setFillColor(colors.HexColor("#777777"))
    canvas.drawCentredString(
        largeur / 2,
        8 * mm,
        f"Registre des stations piscicoles — Page {doc.page}",
    )
    canvas.restoreState()


@router.post("/export/pdf")
def exporter_pdf(req: ExportRequest, db: Session = Depends(get_db)):
    stations = _stations_filtrees(req, db)

    tampon = io.BytesIO()
    doc = SimpleDocTemplate(
        tampon,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=26 * mm,
        bottomMargin=16 * mm,
        title="Registre des stations piscicoles",
    )

    style_cellule = ParagraphStyle(
        "cellule", fontName="Helvetica", fontSize=6.5, leading=8
    )
    style_entete = ParagraphStyle(
        "entete",
        fontName="Helvetica-Bold",
        fontSize=7,
        leading=9,
        textColor=colors.white,
    )
    style_titre = ParagraphStyle(
        "titre", fontName="Helvetica-Bold", fontSize=13, alignment=1, spaceAfter=4
    )
    style_sous_titre = ParagraphStyle(
        "sous_titre",
        fontName="Helvetica",
        fontSize=8,
        alignment=1,
        textColor=colors.HexColor("#555555"),
    )

    colonnes_pdf = [
        "Code",
        "Nom",
        "Province / Localité",
        "Type",
        "Espèces",
        "Promoteur",
        "Capacité\n(t/an)",
        "Agrément",
        "Statut",
    ]

    donnees_table = [
        [Paragraph(c.replace("\n", "<br/>"), style_entete) for c in colonnes_pdf]
    ]
    styles_statut = []

    for i, s in enumerate(stations, start=1):
        localisation = s.province + (f"<br/>{s.localite}" if s.localite else "")
        especes = s.especes_elevees.replace(",", ", ") if s.especes_elevees else "—"
        donnees_table.append(
            [
                Paragraph(s.code_station, style_cellule),
                Paragraph(s.nom, style_cellule),
                Paragraph(localisation, style_cellule),
                Paragraph(
                    TYPE_STATION_LABELS.get(s.type_station.name, ""), style_cellule
                ),
                Paragraph(especes, style_cellule),
                Paragraph(s.promoteur_nom, style_cellule),
                Paragraph(
                    (
                        str(s.capacite_production)
                        if s.capacite_production is not None
                        else "—"
                    ),
                    style_cellule,
                ),
                Paragraph(s.numero_agrement or "Non agréée", style_cellule),
                Paragraph(
                    f"<b>{STATUT_LABELS.get(s.statut.name, '')}</b>", style_cellule
                ),
            ]
        )
        styles_statut.append(
            (
                "TEXTCOLOR",
                (8, i),
                (8, i),
                COULEURS_STATUT_PDF.get(s.statut.name, colors.black),
            )
        )

    largeurs = [
        22 * mm,
        48 * mm,
        30 * mm,
        24 * mm,
        30 * mm,
        40 * mm,
        16 * mm,
        26 * mm,
        24 * mm,
    ]

    table = Table(donnees_table, colWidths=largeurs, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1b5e20")),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [colors.white, colors.HexColor("#f1f8e9")],
                ),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#bdbdbd")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 3),
                ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                ("ALIGN", (6, 1), (6, -1), "RIGHT"),
                *styles_statut,
            ]
        )
    )

    elements = [
        Paragraph("REGISTRE DES STATIONS PISCICOLES", style_titre),
        Paragraph(
            f"{len(stations)} station(s)"
            + (
                f" — filtres : {', '.join(f'{k}={v}' for k, v in req.model_dump(exclude_none=True).items())}"
                if req.model_dump(exclude_none=True)
                else ""
            ),
            style_sous_titre,
        ),
        Spacer(1, 6 * mm),
        table,
    ]

    doc.build(elements, onFirstPage=_entete_pied_pdf, onLaterPages=_entete_pied_pdf)
    return _reponse_fichier(tampon, "application/pdf", "pdf")
