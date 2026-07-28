# -*- coding: utf-8 -*-
"""
Router Import/Export — Module Stations Piscicoles (SIGDP-GABON)
- POST /stations-piscicoles/import          : import de fichier .xlsx, .csv ou .json
- POST /stations-piscicoles/import/modele   : téléchargement du modèle Excel

Normalisation : enums tolérants aux libellés français/accents/casse,
nombres au format français ("1 234,56"), dates dd/mm/yyyy ou ISO.
Dépendance : pip install openpyxl
"""

import csv
import io
import json
import unicodedata
from datetime import datetime, date

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from app.database import get_db  # à ajuster selon l'arborescence
from app.models.stations_piscicole import (
    StationPiscicole,
    TypeStationEnum,
    SourceEauEnum,
    TypePromoteurEnum,
    StatutStationEnum,
)

router = APIRouter(
    prefix="/api/stations-piscicoles", tags=["Stations piscicoles - Import"]
)


# ---------------------------------------------------------------------------
# Normalisation
# ---------------------------------------------------------------------------


def _sans_accents(texte: str) -> str:
    return "".join(
        c
        for c in unicodedata.normalize("NFD", texte)
        if unicodedata.category(c) != "Mn"
    )


def _normaliser_cle(valeur) -> str:
    """'Bacs hors-sol' -> 'BACS_HORS_SOL' ; 'Rivière' -> 'RIVIERE'."""
    if valeur is None:
        return ""
    texte = _sans_accents(str(valeur).strip()).upper()
    for sep in (" ", "-", "'", "/"):
        texte = texte.replace(sep, "_")
    while "__" in texte:
        texte = texte.replace("__", "_")
    return texte.strip("_")


def _parser_enum(valeur, enum_cls, alias: dict = None, obligatoire=False, champ=""):
    """Convertit un libellé libre vers l'enum, avec alias optionnels."""
    cle = _normaliser_cle(valeur)
    if not cle:
        if obligatoire:
            raise ValueError(f"Le champ '{champ}' est obligatoire")
        return None
    if alias and cle in alias:
        cle = alias[cle]
    try:
        return enum_cls[cle]
    except KeyError:
        valides = ", ".join(e.name for e in enum_cls)
        raise ValueError(
            f"Valeur '{valeur}' invalide pour '{champ}' (attendu : {valides})"
        )


ALIAS_TYPE_STATION = {
    "ETANG": "ETANGS",
    "BAC_HORS_SOL": "BACS_HORS_SOL",
    "BACS": "BACS_HORS_SOL",
    "CAGE_FLOTTANTE": "CAGES_FLOTTANTES",
    "CAGES": "CAGES_FLOTTANTES",
}

ALIAS_SOURCE_EAU = {
    "RESEAU_D_EAU": "RESEAU",
    "SEG": "RESEAU",
}

ALIAS_PROMOTEUR = {
    "COOPERATIVE_": "COOPERATIVE",
    "ETAT": "ETATIQUE",
    "PUBLIC": "ETATIQUE",
}


def _parser_nombre(valeur, champ=""):
    """Nombres français : '1 234,56' / '1\u202f234,56' -> 1234.56"""
    if valeur is None or valeur == "":
        return None
    if isinstance(valeur, (int, float)):
        return float(valeur)
    texte = str(valeur).strip()
    for espace in (" ", "\u00a0", "\u202f"):
        texte = texte.replace(espace, "")
    texte = texte.replace(",", ".")
    try:
        return float(texte)
    except ValueError:
        raise ValueError(f"Nombre invalide '{valeur}' pour '{champ}'")


def _parser_date(valeur, champ=""):
    if valeur is None or valeur == "":
        return None
    if isinstance(valeur, datetime):
        return valeur.date()
    if isinstance(valeur, date):
        return valeur
    texte = str(valeur).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(texte, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Date invalide '{valeur}' pour '{champ}' (attendu jj/mm/aaaa)")


def _normaliser_especes(valeur) -> str:
    """'tilapia ; Clarias' -> 'TILAPIA,CLARIAS'"""
    if not valeur:
        return None
    brut = str(valeur).replace(";", ",")
    especes = [_normaliser_cle(e) for e in brut.split(",") if e.strip()]
    return ",".join(especes) if especes else None


# ---------------------------------------------------------------------------
# Colonnes du modèle — l'ordre définit le modèle Excel
# ---------------------------------------------------------------------------

COLONNES = [
    "Nom",
    "Province",
    "Département",
    "Localité",
    "Adresse",
    "Latitude",
    "Longitude",
    "Type de station",
    "Superficie (m²)",
    "Nombre de bassins",
    "Capacité (t/an)",
    "Source d'eau",
    "Espèces",
    "Promoteur",
    "Contact promoteur",
    "Type promoteur",
    "Numéro agrément",
    "Date agrément",
    "Date expiration agrément",
    "Observations",
]

# En-tête normalisé -> nom de colonne canonique (tolère variations d'accents/casse)
_INDEX_COLONNES = {_normaliser_cle(c): c for c in COLONNES}


def _generer_code_station(db: Session) -> str:
    annee = datetime.utcnow().year
    dernier = (
        db.query(StationPiscicole.code_station)
        .filter(StationPiscicole.code_station.like(f"SP-{annee}-%"))
        .order_by(StationPiscicole.code_station.desc())
        .first()
    )
    numero = int(dernier[0].split("-")[-1]) + 1 if dernier else 1
    return f"SP-{annee}-{numero:04d}"


def _ligne_vers_station(ligne: dict, db: Session) -> StationPiscicole:
    """Convertit une ligne normalisée {colonne_canonique: valeur} en modèle."""
    nom = str(ligne.get("Nom") or "").strip()
    if not nom:
        raise ValueError("Le champ 'Nom' est obligatoire")
    province = str(ligne.get("Province") or "").strip()
    if not province:
        raise ValueError("Le champ 'Province' est obligatoire")
    promoteur = str(ligne.get("Promoteur") or "").strip()
    if not promoteur:
        raise ValueError("Le champ 'Promoteur' est obligatoire")

    latitude = _parser_nombre(ligne.get("Latitude"), "Latitude")
    longitude = _parser_nombre(ligne.get("Longitude"), "Longitude")
    if latitude is not None and not -90 <= latitude <= 90:
        raise ValueError(f"Latitude hors bornes : {latitude}")
    if longitude is not None and not -180 <= longitude <= 180:
        raise ValueError(f"Longitude hors bornes : {longitude}")
    # Écarter le point (0,0) — Null Island
    if latitude == 0 and longitude == 0:
        latitude = longitude = None

    return StationPiscicole(
        code_station=_generer_code_station(db),
        nom=nom,
        province=province,
        departement=str(ligne.get("Département") or "").strip() or None,
        localite=str(ligne.get("Localité") or "").strip() or None,
        adresse=str(ligne.get("Adresse") or "").strip() or None,
        latitude=latitude,
        longitude=longitude,
        type_station=_parser_enum(
            ligne.get("Type de station"),
            TypeStationEnum,
            ALIAS_TYPE_STATION,
            obligatoire=True,
            champ="Type de station",
        ),
        superficie_totale=_parser_nombre(ligne.get("Superficie (m²)"), "Superficie"),
        nombre_bassins=int(
            _parser_nombre(ligne.get("Nombre de bassins"), "Nombre de bassins") or 0
        )
        or None,
        capacite_production=_parser_nombre(ligne.get("Capacité (t/an)"), "Capacité"),
        source_eau=_parser_enum(
            ligne.get("Source d'eau"),
            SourceEauEnum,
            ALIAS_SOURCE_EAU,
            champ="Source d'eau",
        ),
        especes_elevees=_normaliser_especes(ligne.get("Espèces")),
        promoteur_nom=promoteur,
        promoteur_contact=str(ligne.get("Contact promoteur") or "").strip() or None,
        promoteur_type=_parser_enum(
            ligne.get("Type promoteur"),
            TypePromoteurEnum,
            ALIAS_PROMOTEUR,
            champ="Type promoteur",
        )
        or TypePromoteurEnum.PRIVE,
        statut=StatutStationEnum.EN_CONSTRUCTION,
        numero_agrement=str(ligne.get("Numéro agrément") or "").strip() or None,
        date_agrement=_parser_date(ligne.get("Date agrément"), "Date agrément"),
        date_expiration_agrement=_parser_date(
            ligne.get("Date expiration agrément"), "Date expiration agrément"
        ),
        observations=str(ligne.get("Observations") or "").strip() or None,
    )


# ---------------------------------------------------------------------------
# Lecture des fichiers
# ---------------------------------------------------------------------------


def _lire_xlsx(contenu: bytes) -> list[dict]:
    wb = load_workbook(io.BytesIO(contenu), data_only=True)
    ws = wb.active
    lignes_iter = ws.iter_rows(values_only=True)
    entetes_bruts = next(lignes_iter, None)
    if not entetes_bruts:
        raise ValueError("Fichier Excel vide")
    entetes = [_INDEX_COLONNES.get(_normaliser_cle(e)) for e in entetes_bruts]
    lignes = []
    for valeurs in lignes_iter:
        if all(v is None or str(v).strip() == "" for v in valeurs):
            continue  # ignorer les lignes vides
        lignes.append({col: val for col, val in zip(entetes, valeurs) if col})
    return lignes


def _lire_csv(contenu: bytes) -> list[dict]:
    texte = contenu.decode("utf-8-sig")  # gère le BOM d'Excel
    # Détection du séparateur (Excel FR exporte en ';')
    try:
        dialecte = csv.Sniffer().sniff(texte[:2048], delimiters=";,\t")
    except csv.Error:
        dialecte = csv.excel
        dialecte.delimiter = ";"
    lecteur = csv.DictReader(io.StringIO(texte), dialect=dialecte)
    lignes = []
    for rangee in lecteur:
        if not any((v or "").strip() for v in rangee.values()):
            continue
        lignes.append(
            {
                _INDEX_COLONNES.get(_normaliser_cle(k)): v
                for k, v in rangee.items()
                if _INDEX_COLONNES.get(_normaliser_cle(k))
            }
        )
    return lignes


def _lire_json(contenu: bytes) -> list[dict]:
    donnees = json.loads(contenu.decode("utf-8"))
    if isinstance(donnees, dict):
        donnees = donnees.get("stations", donnees.get("items", []))
    if not isinstance(donnees, list):
        raise ValueError("Le JSON doit contenir une liste de stations")
    lignes = []
    for objet in donnees:
        lignes.append(
            {
                _INDEX_COLONNES.get(_normaliser_cle(k), None) or k: v
                for k, v in objet.items()
            }
        )
    return lignes


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------


@router.post("/import")
async def importer_stations(
    fichier: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """
    Importe des stations depuis un fichier .xlsx, .csv ou .json.
    Retourne un rapport : lignes importées, lignes en erreur avec le motif.
    Les lignes valides sont insérées même si d'autres lignes échouent.
    """
    nom_fichier = (fichier.filename or "").lower()
    contenu = await fichier.read()
    if len(contenu) > 10 * 1024 * 1024:
        raise HTTPException(
            status_code=400, detail="Fichier trop volumineux (max 10 Mo)"
        )

    try:
        if nom_fichier.endswith(".xlsx"):
            lignes = _lire_xlsx(contenu)
        elif nom_fichier.endswith(".csv") or nom_fichier.endswith(".txt"):
            lignes = _lire_csv(contenu)
        elif nom_fichier.endswith(".json"):
            lignes = _lire_json(contenu)
        else:
            raise HTTPException(
                status_code=400,
                detail="Format non supporté (attendu : .xlsx, .csv, .txt ou .json)",
            )
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Fichier illisible : {exc}")

    if not lignes:
        raise HTTPException(status_code=400, detail="Aucune ligne de données trouvée")

    importees = 0
    erreurs = []

    for index, ligne in enumerate(lignes, start=2):  # ligne 1 = en-têtes
        try:
            # Savepoint par ligne : une erreur n'annule que cette ligne,
            # pas les lignes valides déjà flushées dans la transaction.
            with db.begin_nested():
                station = _ligne_vers_station(ligne, db)
                db.add(station)
                db.flush()  # détecte immédiatement les contraintes (agrément dupliqué...)
            importees += 1
        except Exception as exc:
            erreurs.append(
                {
                    "ligne": index,
                    "nom": str(ligne.get("Nom") or "")[:60],
                    "erreur": str(exc),
                }
            )

    db.commit()

    return {
        "total_lignes": len(lignes),
        "importees": importees,
        "rejetees": len(erreurs),
        "erreurs": erreurs,
    }


@router.post("/import/modele")
def telecharger_modele():
    """Génère le modèle Excel d'import avec en-têtes stylés et ligne d'exemple."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Stations piscicoles"

    # En-têtes
    entete_fill = PatternFill("solid", start_color="1B5E20")
    entete_font = Font(bold=True, color="FFFFFF")
    for col, nom in enumerate(COLONNES, start=1):
        cellule = ws.cell(row=1, column=col, value=nom)
        cellule.fill = entete_fill
        cellule.font = entete_font
        ws.column_dimensions[cellule.column_letter].width = max(len(nom) + 4, 14)

    # Ligne d'exemple
    exemple = [
        "Station Piscicole de Ntoum",
        "Estuaire",
        "Komo-Mondah",
        "Ntoum",
        "BP 123, Ntoum",
        "0,3901",
        "9,7671",
        "Étangs",
        "12 000",
        "14",
        "25,5",
        "Rivière",
        "TILAPIA, CLARIAS",
        "Coopérative Aquacole de Ntoum",
        "+241 66 00 00 01",
        "Coopérative",
        "AGR-SP-2026-001",
        "15/03/2026",
        "15/03/2031",
        "Exemple — supprimer cette ligne avant import",
    ]
    for col, valeur in enumerate(exemple, start=1):
        ws.cell(row=2, column=col, value=valeur)

    tampon = io.BytesIO()
    wb.save(tampon)
    tampon.seek(0)

    return StreamingResponse(
        tampon,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="modele_import_stations_piscicoles.xlsx"'
        },
    )
