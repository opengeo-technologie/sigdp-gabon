from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, extract
from datetime import datetime, date
from typing import Optional
import io
import csv
import json

from app.database import get_db
from app.models.debarcadere import Debarcadere
from app.models.pecheur import Pecheur
from app.models.bateau import Bateau
from app.models.espece import Espece
from app.models.debarquement import Debarquement, DetailDebarquement

router = APIRouter(prefix="/api/rapports", tags=["Rapports"])


def generer_csv(data: list, headers: list) -> io.StringIO:
    """Génère un fichier CSV à partir de données"""
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=headers)
    writer.writeheader()
    writer.writerows(data)
    output.seek(0)
    return output


def generer_excel(data: list, headers: list, titre: str) -> io.BytesIO:
    """Génère un fichier Excel à partir de données"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = titre[:31]  # Excel limite à 31 caractères

        # En-tête
        header_fill = PatternFill(
            start_color="2196F3", end_color="2196F3", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Données
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))

        # Ajuster la largeur des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl non installé")


def generer_pdf(data: list, headers: list, titre: str) -> io.BytesIO:
    """Génère un fichier PDF à partir de données"""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import (
            SimpleDocTemplate,
            Table,
            TableStyle,
            Paragraph,
            Spacer,
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib.enums import TA_CENTER
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont

        output = io.BytesIO()

        # Utiliser paysage si beaucoup de colonnes
        pagesize = landscape(A4) if len(headers) > 6 else A4
        doc = SimpleDocTemplate(
            output, pagesize=pagesize, topMargin=0.5 * inch, bottomMargin=0.5 * inch
        )

        elements = []
        styles = getSampleStyleSheet()

        # Titre
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=18,
            textColor=colors.HexColor("#1976d2"),
            spaceAfter=30,
            alignment=TA_CENTER,
        )

        elements.append(Paragraph(titre, title_style))
        elements.append(Spacer(1, 0.3 * inch))

        # Date de génération
        date_style = ParagraphStyle(
            "DateStyle",
            parent=styles["Normal"],
            fontSize=10,
            textColor=colors.grey,
            alignment=TA_CENTER,
        )
        elements.append(
            Paragraph(
                f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", date_style
            )
        )
        elements.append(Spacer(1, 0.3 * inch))

        # Préparer les données du tableau
        table_data = [headers]
        for row in data:
            table_data.append([str(row.get(h, "")) for h in headers])

        # Créer le tableau
        table = Table(table_data)

        # Style du tableau
        table.setStyle(
            TableStyle(
                [
                    # En-tête
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2196F3")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    # Données
                    ("ALIGN", (0, 1), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                    ("FONTSIZE", (0, 1), (-1, -1), 8),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [colors.white, colors.HexColor("#f5f5f5")],
                    ),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )

        elements.append(table)

        # Générer le PDF
        doc.build(elements)
        output.seek(0)
        return output

    except ImportError as e:
        raise HTTPException(
            status_code=500, detail=f"Bibliothèque PDF manquante: {str(e)}"
        )
        elements.append(Paragraph(titre, title_style))

        # Date de génération
        date_style = ParagraphStyle(
            "DateStyle",
            parent=styles["Normal"],
            fontSize=10,
            textColor=colors.grey,
            alignment=TA_CENTER,
            spaceAfter=20,
        )
        elements.append(
            Paragraph(
                f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", date_style
            )
        )
        elements.append(Spacer(1, 0.2 * inch))

        # Préparer les données du tableau
        table_data = [headers]
        for row in data:
            table_data.append([str(row.get(h, "")) for h in headers])

        # Limiter le nombre de lignes par page
        max_rows_per_page = 40

        for i in range(0, len(table_data), max_rows_per_page):
            chunk = table_data[i : min(i + max_rows_per_page, len(table_data))]

            # Si ce n'est pas la première page, ajouter les en-têtes
            if i > 0:
                chunk = [headers] + chunk

            # Créer le tableau
            table = Table(chunk)

            # Style du tableau
            style = TableStyle(
                [
                    # En-tête
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2196F3")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    # Corps
                    ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                    ("FONTSIZE", (0, 1), (-1, -1), 8),
                    ("ALIGN", (0, 1), (-1, -1), "LEFT"),
                    # Grille
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    # Alternance de couleurs
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [colors.white, colors.HexColor("#f5f5f5")],
                    ),
                ]
            )

            table.setStyle(style)
            elements.append(table)

            # Ajouter un saut de page si ce n'est pas la dernière page
            if i + max_rows_per_page < len(table_data):
                elements.append(PageBreak())

        # Générer le PDF
        doc.build(elements)
        output.seek(0)
        return output

    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="reportlab non installé. Installez avec: pip install reportlab",
        )


@router.post("/generer")
def generer_rapport(
    type: str = Query(...),
    format: str = Query("pdf", regex="^(pdf|excel|csv)$"),
    date_debut: Optional[date] = None,
    date_fin: Optional[date] = None,
    debarcadere_id: Optional[int] = None,
    espece_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    """
    Génère un rapport selon le type demandé
    """

    # Par défaut : dernier mois
    if not date_fin:
        date_fin = datetime.now().date()
    if not date_debut:
        date_debut = date(date_fin.year, date_fin.month, 1)

    # Dispatcher selon le type de rapport
    if type == "debarquements":
        data, headers = rapport_debarquements(db, date_debut, date_fin, debarcadere_id)
        titre = "Débarquements"

    elif type == "captures_par_espece":
        data, headers = rapport_captures_par_espece(db, date_debut, date_fin, espece_id)
        titre = "Captures par espèce"

    elif type == "flotte_bateaux":
        data, headers = rapport_flotte_bateaux(db)
        titre = "Flotte de bateaux"

    elif type == "pecheurs_actifs":
        data, headers = rapport_pecheurs_actifs(db, date_debut, date_fin)
        titre = "Pêcheurs actifs"

    elif type == "activite_debarcaderes":
        data, headers = rapport_activite_debarcaderes(db, date_debut, date_fin)
        titre = "Activité débarcadères"

    elif type == "quotas_utilisation":
        data, headers = rapport_quotas_utilisation(db, date_debut, date_fin)
        titre = "Utilisation quotas"

    elif type == "alertes":
        data, headers = rapport_alertes(db, date_debut, date_fin)
        titre = "Alertes"

    elif type == "valeur_economique":
        data, headers = rapport_valeur_economique(db, date_debut, date_fin)
        titre = "Valeur économique"

    elif type == "synthese_activite":
        data, headers = rapport_synthese_activite(db, date_debut, date_fin)
        titre = "Synthèse d'activité"

    elif type == "conformite":
        data, headers = rapport_conformite(db)
        titre = "Conformité"

    # NOUVEAUX RAPPORTS
    elif type == "production_mensuelle":
        data, headers = rapport_production_mensuelle(db, date_debut, date_fin)
        titre = "Production mensuelle"

    elif type == "engins_peche":
        data, headers = rapport_engins_peche(db, date_debut, date_fin)
        titre = "Utilisation des engins"

    elif type == "zones_peche":
        data, headers = rapport_zones_peche(db, date_debut, date_fin)
        titre = "Activité par zone"

    elif type == "effort_peche":
        data, headers = rapport_effort_peche(db, date_debut, date_fin)
        titre = "Effort de pêche"

    elif type == "saisonnalite":
        data, headers = rapport_saisonnalite(db, date_debut, date_fin)
        titre = "Saisonnalité"

    elif type == "tailles_captures":
        data, headers = rapport_tailles_captures(db, date_debut, date_fin, espece_id)
        titre = "Distribution des tailles"

    elif type == "prix_marche":
        data, headers = rapport_prix_marche(db, date_debut, date_fin)
        titre = "Évolution des prix"

    elif type == "rentabilite_sortie":
        data, headers = rapport_rentabilite_sortie(db, date_debut, date_fin)
        titre = "Rentabilité par sortie"

    elif type == "composition_captures":
        data, headers = rapport_composition_captures(db, date_debut, date_fin)
        titre = "Composition des captures"

    elif type == "surveillance_ressources":
        data, headers = rapport_surveillance_ressources(db, date_debut, date_fin)
        titre = "Surveillance des ressources"

    else:
        raise HTTPException(status_code=400, detail="Type de rapport inconnu")

    # Générer selon le format
    if format == "csv":
        output = generer_csv(data, headers)
        filename = f"rapport_{type}_{date_debut}_{date_fin}.csv"
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode("utf-8-sig")),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    elif format == "excel":
        output = generer_excel(data, headers, titre)
        filename = f"rapport_{type}_{date_debut}_{date_fin}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    elif format == "pdf":
        output = generer_pdf(data, headers, titre)
        filename = f"rapport_{type}_{date_debut}_{date_fin}.pdf"
        return StreamingResponse(
            output,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )


def rapport_debarquements(
    db: Session, date_debut: date, date_fin: date, debarcadere_id: Optional[int] = None
):
    """Rapport détaillé des débarquements"""
    query = (
        db.query(
            Debarquement.id,
            Debarquement.date_debarquement,
            Debarcadere.denomination.label("debarcadere"),
            Pecheur.nom.label("pecheur_nom"),
            Pecheur.prenom.label("pecheur_prenom"),
            Bateau.numero_immatriculation.label("bateau"),
            Debarquement.heure_depart_peche,
            Debarquement.heure_arrivee_debarcadere,
            func.sum(DetailDebarquement.quantite_kg).label("quantite_totale_kg"),
            func.sum(DetailDebarquement.valeur_totale).label("valeur_totale_fcfa"),
            func.count(DetailDebarquement.id).label("nombre_especes"),
        )
        .select_from(Debarquement)
        .join(Debarcadere, Debarquement.debarcadere_id == Debarcadere.id)
        .join(Pecheur, Debarquement.pecheur_principal_id == Pecheur.id)
        .join(Bateau, Debarquement.bateau_id == Bateau.id)
        .join(DetailDebarquement, DetailDebarquement.debarquement_id == Debarquement.id)
        .filter(Debarquement.date_debarquement.between(date_debut, date_fin))
    )

    if debarcadere_id:
        query = query.filter(Debarquement.debarcadere_id == debarcadere_id)

    query = query.group_by(
        Debarquement.id,
        Debarquement.date_debarquement,
        Debarcadere.denomination,
        Pecheur.nom,
        Pecheur.prenom,
        Bateau.numero_immatriculation,
        Debarquement.heure_depart_peche,
        Debarquement.heure_arrivee_debarcadere,
    ).order_by(Debarquement.date_debarquement.desc())

    resultats = query.all()

    data = []
    for r in resultats:
        data.append(
            {
                "ID": r.id,
                "Date": r.date_debarquement.strftime("%Y-%m-%d"),
                "Débarcadère": r.debarcadere,
                "Pêcheur": f"{r.pecheur_nom} {r.pecheur_prenom}",
                "Bateau": r.bateau,
                "Heure départ": (
                    r.heure_depart_peche.strftime("%H:%M")
                    if r.heure_depart_peche
                    else ""
                ),
                "Heure arrivée": (
                    r.heure_arrivee_debarcadere.strftime("%H:%M")
                    if r.heure_arrivee_debarcadere
                    else ""
                ),
                "Quantité (kg)": round(float(r.quantite_totale_kg or 0), 2),
                "Valeur (FCFA)": round(float(r.valeur_totale_fcfa or 0), 2),
                "Nb espèces": r.nombre_especes,
            }
        )

    headers = [
        "ID",
        "Date",
        "Débarcadère",
        "Pêcheur",
        "Bateau",
        "Heure départ",
        "Heure arrivée",
        "Quantité (kg)",
        "Valeur (FCFA)",
        "Nb espèces",
    ]

    return data, headers


def rapport_captures_par_espece(
    db: Session, date_debut: date, date_fin: date, espece_id: Optional[int] = None
):
    """Rapport des captures groupées par espèce"""
    query = (
        db.query(
            Espece.nom_commun_francais,
            Espece.nom_scientifique,
            Espece.code_espece,
            func.sum(DetailDebarquement.quantite_kg).label("quantite_totale_kg"),
            func.sum(DetailDebarquement.valeur_totale).label("valeur_totale_fcfa"),
            func.count(DetailDebarquement.id).label("nombre_captures"),
            func.avg(DetailDebarquement.prix_unitaire_kg).label("prix_moyen_kg"),
        )
        .select_from(Espece)
        .join(DetailDebarquement, DetailDebarquement.espece_id == Espece.id)
        .join(Debarquement, Debarquement.id == DetailDebarquement.debarquement_id)
        .filter(Debarquement.date_debarquement.between(date_debut, date_fin))
    )

    if espece_id:
        query = query.filter(Espece.id == espece_id)

    query = query.group_by(
        Espece.id,
        Espece.nom_commun_francais,
        Espece.nom_scientifique,
        Espece.code_espece,
    ).order_by(func.sum(DetailDebarquement.quantite_kg).desc())

    resultats = query.all()

    data = []
    for r in resultats:
        data.append(
            {
                "Espèce": r.nom_commun_francais,
                "Nom scientifique": r.nom_scientifique or "",
                "Code": r.code_espece or "",
                "Quantité totale (kg)": round(float(r.quantite_totale_kg or 0), 2),
                "Quantité (tonnes)": round(float(r.quantite_totale_kg or 0) / 1000, 3),
                "Valeur totale (FCFA)": round(float(r.valeur_totale_fcfa or 0), 2),
                "Nombre de captures": r.nombre_captures,
                "Prix moyen (FCFA/kg)": round(float(r.prix_moyen_kg or 0), 2),
            }
        )

    headers = [
        "Espèce",
        "Nom scientifique",
        "Code",
        "Quantité totale (kg)",
        "Quantité (tonnes)",
        "Valeur totale (FCFA)",
        "Nombre de captures",
        "Prix moyen (FCFA/kg)",
    ]

    return data, headers


def rapport_flotte_bateaux(db: Session):
    """Rapport de la flotte de bateaux"""
    bateaux = db.query(Bateau).all()

    data = []
    for b in bateaux:
        data.append(
            {
                "Immatriculation": b.numero_immatriculation,
                "Nom": b.nom_bateau or "",
                "Type": b.type_bateau,
                "Propulsion": b.propulsion,
                "Longueur (m)": b.longueur_hors_tout or "",
                "Matériau": b.materiau_coque,
                "Puissance (CV)": b.moteur_puissance_cv or "",
                "Propriétaire": b.proprietaire_nom or "",
                "Statut": b.statut,
            }
        )

    headers = [
        "Immatriculation",
        "Nom",
        "Type",
        "Propulsion",
        "Longueur (m)",
        "Matériau",
        "Puissance (CV)",
        "Propriétaire",
        "Statut",
    ]

    return data, headers


def rapport_pecheurs_actifs(db: Session, date_debut: date, date_fin: date):
    """Rapport des pêcheurs actifs"""
    pecheurs = (
        db.query(
            Pecheur.numero_carte,
            Pecheur.nom,
            Pecheur.prenom,
            Pecheur.categorie,
            Pecheur.type_peche,
            func.count(Debarquement.id).label("nb_sorties"),
            func.sum(DetailDebarquement.quantite_kg).label("total_kg"),
        )
        .select_from(Pecheur)
        .join(Debarquement, Debarquement.pecheur_principal_id == Pecheur.id)
        .join(DetailDebarquement, DetailDebarquement.debarquement_id == Debarquement.id)
        .filter(Debarquement.date_debarquement.between(date_debut, date_fin))
        .group_by(
            Pecheur.id,
            Pecheur.numero_carte,
            Pecheur.nom,
            Pecheur.prenom,
            Pecheur.categorie,
            Pecheur.type_peche,
        )
        .order_by(func.count(Debarquement.id).desc())
        .all()
    )

    data = []
    for p in pecheurs:
        data.append(
            {
                "N° Carte": p.numero_carte,
                "Nom": p.nom,
                "Prénom": p.prenom,
                "Catégorie": p.categorie,
                "Type pêche": p.type_peche,
                "Nombre sorties": p.nb_sorties,
                "Total captures (kg)": round(float(p.total_kg or 0), 2),
            }
        )

    headers = [
        "N° Carte",
        "Nom",
        "Prénom",
        "Catégorie",
        "Type pêche",
        "Nombre sorties",
        "Total captures (kg)",
    ]

    return data, headers


def rapport_activite_debarcaderes(db: Session, date_debut: date, date_fin: date):
    """Rapport d'activité par débarcadère"""
    resultats = (
        db.query(
            Debarcadere.denomination,
            Debarcadere.code,
            Debarcadere.province,
            func.count(Debarquement.id).label("nb_debarquements"),
            func.sum(DetailDebarquement.quantite_kg).label("total_kg"),
        )
        .select_from(Debarcadere)
        .join(Debarquement, Debarquement.debarcadere_id == Debarcadere.id)
        .join(DetailDebarquement, DetailDebarquement.debarquement_id == Debarquement.id)
        .filter(Debarquement.date_debarquement.between(date_debut, date_fin))
        .group_by(
            Debarcadere.id,
            Debarcadere.denomination,
            Debarcadere.code,
            Debarcadere.province,
        )
        .order_by(func.count(Debarquement.id).desc())
        .all()
    )

    data = []
    for r in resultats:
        data.append(
            {
                "Débarcadère": r.denomination,
                "Code": r.code,
                "Province": r.province,
                "Nb débarquements": r.nb_debarquements,
                "Total captures (kg)": round(float(r.total_kg or 0), 2),
                "Total (tonnes)": round(float(r.total_kg or 0) / 1000, 3),
            }
        )

    headers = [
        "Débarcadère",
        "Code",
        "Province",
        "Nb débarquements",
        "Total captures (kg)",
        "Total (tonnes)",
    ]

    return data, headers


def rapport_quotas_utilisation(db: Session, date_debut: date, date_fin: date):
    """Rapport d'utilisation des quotas"""
    resultats = (
        db.query(
            Espece.nom_commun_francais,
            Espece.quota_annuel,
            func.sum(DetailDebarquement.quantite_kg).label("quantite_capturee"),
            (
                func.sum(DetailDebarquement.quantite_kg) / Espece.quota_annuel * 100
            ).label("pourcentage_utilise"),
        )
        .select_from(Espece)
        .join(DetailDebarquement, DetailDebarquement.espece_id == Espece.id)
        .join(Debarquement, Debarquement.id == DetailDebarquement.debarquement_id)
        .filter(Debarquement.date_debarquement.between(date_debut, date_fin))
        .filter(Espece.quota_annuel.isnot(None))
        .filter(Espece.quota_annuel > 0)
        .group_by(Espece.id, Espece.nom_commun_francais, Espece.quota_annuel)
        .order_by(
            (func.sum(DetailDebarquement.quantite_kg) / Espece.quota_annuel).desc()
        )
        .all()
    )

    data = []
    for r in resultats:
        quota_kg = float(r.quota_annuel or 0)
        capture_kg = float(r.quantite_capturee or 0)
        pourcentage = (capture_kg / quota_kg * 100) if quota_kg > 0 else 0

        statut = "Normal"
        if pourcentage >= 100:
            statut = "Quota dépassé"
        elif pourcentage >= 90:
            statut = "Alerte"
        elif pourcentage >= 75:
            statut = "Surveillance"

        data.append(
            {
                "Espèce": r.nom_commun_francais,
                "Quota annuel (kg)": round(quota_kg, 2),
                "Quota annuel (tonnes)": round(quota_kg / 1000, 3),
                "Capturé (kg)": round(capture_kg, 2),
                "Capturé (tonnes)": round(capture_kg / 1000, 3),
                "% Utilisé": round(pourcentage, 2),
                "Statut": statut,
            }
        )

    headers = [
        "Espèce",
        "Quota annuel (kg)",
        "Quota annuel (tonnes)",
        "Capturé (kg)",
        "Capturé (tonnes)",
        "% Utilisé",
        "Statut",
    ]

    return data, headers


def rapport_alertes(db: Session, date_debut: date, date_fin: date):
    """Rapport des alertes"""
    # Récupérer les débarquements avec alertes
    debarquements = (
        db.query(
            Debarquement.id,
            Debarquement.date_debarquement,
            Debarquement.alertes,
            Debarcadere.denomination.label("debarcadere"),
            Pecheur.nom.label("pecheur_nom"),
            Pecheur.prenom.label("pecheur_prenom"),
        )
        .select_from(Debarquement)
        .join(Debarcadere, Debarquement.debarcadere_id == Debarcadere.id)
        .join(Pecheur, Debarquement.pecheur_principal_id == Pecheur.id)
        .filter(Debarquement.date_debarquement.between(date_debut, date_fin))
        .filter(Debarquement.alertes.isnot(None))
        .filter(Debarquement.alertes != "")
        .order_by(Debarquement.date_debarquement.desc())
        .all()
    )

    data = []
    for d in debarquements:
        # Parser les alertes (format JSON ou texte)
        alertes_list = d.alertes.split(";") if d.alertes else []

        for alerte in alertes_list:
            if alerte.strip():
                # Déterminer le type d'alerte
                type_alerte = "Info"
                if "quota" in alerte.lower() or "dépassé" in alerte.lower():
                    type_alerte = "Quota dépassé"
                elif "protégée" in alerte.lower() or "interdite" in alerte.lower():
                    type_alerte = "Espèce protégée"
                elif "taille" in alerte.lower() or "minimal" in alerte.lower():
                    type_alerte = "Taille minimale"
                elif "certificat" in alerte.lower() or "licence" in alerte.lower():
                    type_alerte = "Conformité"

                data.append(
                    {
                        "Date": d.date_debarquement.strftime("%Y-%m-%d"),
                        "ID Débarquement": d.id,
                        "Type": type_alerte,
                        "Description": alerte.strip(),
                        "Débarcadère": d.debarcadere,
                        "Pêcheur": f"{d.pecheur_nom} {d.pecheur_prenom}",
                    }
                )

    headers = [
        "Date",
        "ID Débarquement",
        "Type",
        "Description",
        "Débarcadère",
        "Pêcheur",
    ]

    return data, headers


def rapport_valeur_economique(db: Session, date_debut: date, date_fin: date):
    """Rapport de valeur économique par mois"""
    # Grouper par mois
    resultats = (
        db.query(
            extract("year", Debarquement.date_debarquement).label("annee"),
            extract("month", Debarquement.date_debarquement).label("mois"),
            func.sum(DetailDebarquement.valeur_totale).label("valeur_totale"),
            func.sum(DetailDebarquement.quantite_kg).label("quantite_totale"),
            func.count(Debarquement.id).label("nb_debarquements"),
        )
        .select_from(Debarquement)
        .join(DetailDebarquement, DetailDebarquement.debarquement_id == Debarquement.id)
        .filter(Debarquement.date_debarquement.between(date_debut, date_fin))
        .group_by("annee", "mois")
        .order_by("annee", "mois")
        .all()
    )

    mois_noms = [
        "",
        "Janvier",
        "Février",
        "Mars",
        "Avril",
        "Mai",
        "Juin",
        "Juillet",
        "Août",
        "Septembre",
        "Octobre",
        "Novembre",
        "Décembre",
    ]

    data = []
    for r in resultats:
        valeur = float(r.valeur_totale or 0)
        quantite = float(r.quantite_totale or 0)
        prix_moyen = (valeur / quantite) if quantite > 0 else 0

        data.append(
            {
                "Période": f"{mois_noms[int(r.mois)]} {int(r.annee)}",
                "Année": int(r.annee),
                "Mois": int(r.mois),
                "Valeur totale (FCFA)": round(valeur, 2),
                "Valeur (Millions FCFA)": round(valeur / 1000000, 3),
                "Quantité (kg)": round(quantite, 2),
                "Quantité (tonnes)": round(quantite / 1000, 3),
                "Prix moyen (FCFA/kg)": round(prix_moyen, 2),
                "Nb débarquements": r.nb_debarquements,
            }
        )

    headers = [
        "Période",
        "Année",
        "Mois",
        "Valeur totale (FCFA)",
        "Valeur (Millions FCFA)",
        "Quantité (kg)",
        "Quantité (tonnes)",
        "Prix moyen (FCFA/kg)",
        "Nb débarquements",
    ]

    return data, headers


def rapport_synthese_activite(db: Session, date_debut: date, date_fin: date):
    """Rapport de synthèse d'activité globale"""
    # Calculer les indicateurs clés

    # Nombre total de débarquements
    nb_debarquements = (
        db.query(func.count(Debarquement.id))
        .filter(Debarquement.date_debarquement.between(date_debut, date_fin))
        .scalar()
        or 0
    )

    # Quantité et valeur totales
    totaux = (
        db.query(
            func.sum(DetailDebarquement.quantite_kg).label("quantite"),
            func.sum(DetailDebarquement.valeur_totale).label("valeur"),
        )
        .select_from(DetailDebarquement)
        .join(Debarquement, Debarquement.id == DetailDebarquement.debarquement_id)
        .filter(Debarquement.date_debarquement.between(date_debut, date_fin))
        .first()
    )

    quantite_totale = float(totaux.quantite or 0)
    valeur_totale = float(totaux.valeur or 0)

    # Nombre de pêcheurs actifs
    nb_pecheurs = (
        db.query(func.count(func.distinct(Debarquement.pecheur_principal_id)))
        .filter(Debarquement.date_debarquement.between(date_debut, date_fin))
        .scalar()
        or 0
    )

    # Nombre de bateaux utilisés
    nb_bateaux = (
        db.query(func.count(func.distinct(Debarquement.bateau_id)))
        .filter(Debarquement.date_debarquement.between(date_debut, date_fin))
        .scalar()
        or 0
    )

    # Nombre d'espèces capturées
    nb_especes = (
        db.query(func.count(func.distinct(DetailDebarquement.espece_id)))
        .join(Debarquement, Debarquement.id == DetailDebarquement.debarquement_id)
        .filter(Debarquement.date_debarquement.between(date_debut, date_fin))
        .scalar()
        or 0
    )

    # Espèce la plus capturée
    espece_top = (
        db.query(
            Espece.nom_commun_francais,
            func.sum(DetailDebarquement.quantite_kg).label("total"),
        )
        .select_from(Espece)
        .join(DetailDebarquement, DetailDebarquement.espece_id == Espece.id)
        .join(Debarquement, Debarquement.id == DetailDebarquement.debarquement_id)
        .filter(Debarquement.date_debarquement.between(date_debut, date_fin))
        .group_by(Espece.id, Espece.nom_commun_francais)
        .order_by(func.sum(DetailDebarquement.quantite_kg).desc())
        .first()
    )

    espece_principale = espece_top.nom_commun_francais if espece_top else "N/A"

    # Débarcadère le plus actif
    debarcadere_top = (
        db.query(Debarcadere.denomination, func.count(Debarquement.id).label("total"))
        .select_from(Debarcadere)
        .join(Debarquement, Debarquement.debarcadere_id == Debarcadere.id)
        .filter(Debarquement.date_debarquement.between(date_debut, date_fin))
        .group_by(Debarcadere.id, Debarcadere.denomination)
        .order_by(func.count(Debarquement.id).desc())
        .first()
    )

    debarcadere_principal = debarcadere_top.denomination if debarcadere_top else "N/A"

    # Moyennes
    quantite_moyenne = quantite_totale / nb_debarquements if nb_debarquements > 0 else 0
    valeur_moyenne = valeur_totale / nb_debarquements if nb_debarquements > 0 else 0

    data = [
        {"Indicateur": "Période", "Valeur": f"{date_debut} au {date_fin}"},
        {"Indicateur": "Nombre de débarquements", "Valeur": str(nb_debarquements)},
        {
            "Indicateur": "Quantité totale capturée (kg)",
            "Valeur": f"{round(quantite_totale, 2):,}",
        },
        {
            "Indicateur": "Quantité totale (tonnes)",
            "Valeur": f"{round(quantite_totale/1000, 3):,}",
        },
        {
            "Indicateur": "Valeur économique totale (FCFA)",
            "Valeur": f"{round(valeur_totale, 2):,}",
        },
        {
            "Indicateur": "Valeur économique (Millions FCFA)",
            "Valeur": f"{round(valeur_totale/1000000, 3):,}",
        },
        {"Indicateur": "Nombre de pêcheurs actifs", "Valeur": str(nb_pecheurs)},
        {"Indicateur": "Nombre de bateaux utilisés", "Valeur": str(nb_bateaux)},
        {"Indicateur": "Nombre d'espèces capturées", "Valeur": str(nb_especes)},
        {
            "Indicateur": "Quantité moyenne par débarquement (kg)",
            "Valeur": f"{round(quantite_moyenne, 2):,}",
        },
        {
            "Indicateur": "Valeur moyenne par débarquement (FCFA)",
            "Valeur": f"{round(valeur_moyenne, 2):,}",
        },
        {"Indicateur": "Espèce la plus capturée", "Valeur": espece_principale},
        {"Indicateur": "Débarcadère le plus actif", "Valeur": debarcadere_principal},
    ]

    headers = ["Indicateur", "Valeur"]

    return data, headers


def rapport_conformite(db: Session):
    """Rapport de conformité réglementaire"""
    from datetime import timedelta

    aujourdhui = datetime.now().date()
    dans_30_jours = aujourdhui + timedelta(days=30)

    data = []

    # Vérifier les certificats de navigabilité des bateaux
    bateaux = db.query(Bateau).filter(Bateau.statut == "Actif").all()

    for bateau in bateaux:
        if bateau.certificat_navigabilite_date_expiration:
            if bateau.certificat_navigabilite_date_expiration < aujourdhui:
                statut = "❌ Expiré"
            elif bateau.certificat_navigabilite_date_expiration <= dans_30_jours:
                statut = "⚠️ Expire bientôt"
            else:
                statut = "✅ Valide"

            data.append(
                {
                    "Type": "Certificat navigabilité",
                    "Élément": f"Bateau {bateau.numero_immatriculation}",
                    "Statut": statut,
                    "Date expiration": bateau.certificat_navigabilite_date_expiration.strftime(
                        "%Y-%m-%d"
                    ),
                    "Jours restants": (
                        bateau.certificat_navigabilite_date_expiration - aujourdhui
                    ).days,
                }
            )

    # Vérifier les licences de pêcheurs
    pecheurs = db.query(Pecheur).all()

    for pecheur in pecheurs:
        if pecheur.date_expiration_carte:
            if pecheur.date_expiration_carte < aujourdhui:
                statut = "❌ Expirée"
            elif pecheur.date_expiration_carte <= dans_30_jours:
                statut = "⚠️ Expire bientôt"
            else:
                statut = "✅ Valide"

            data.append(
                {
                    "Type": "Carte de pêcheur",
                    "Élément": f"{pecheur.nom} {pecheur.prenom} ({pecheur.numero_carte})",
                    "Statut": statut,
                    "Date expiration": pecheur.date_expiration_carte.strftime(
                        "%Y-%m-%d"
                    ),
                    "Jours restants": (pecheur.date_expiration_carte - aujourdhui).days,
                }
            )

    # Trier par date d'expiration
    data.sort(key=lambda x: x["Jours restants"])

    headers = ["Type", "Élément", "Statut", "Date expiration", "Jours restants"]

    return data, headers


# ========== NOUVEAUX RAPPORTS ==========


def rapport_production_mensuelle(db: Session, date_debut: date, date_fin: date):
    """Rapport de production halieutique mensuelle"""
    resultats = (
        db.query(
            extract("year", Debarquement.date_debarquement).label("annee"),
            extract("month", Debarquement.date_debarquement).label("mois"),
            func.sum(DetailDebarquement.quantite_kg).label("quantite"),
            func.sum(DetailDebarquement.valeur_totale).label("valeur"),
            func.count(Debarquement.id).label("nb_debarquements"),
            func.count(func.distinct(Debarquement.pecheur_principal_id)).label(
                "nb_pecheurs"
            ),
        )
        .select_from(Debarquement)
        .join(DetailDebarquement, DetailDebarquement.debarquement_id == Debarquement.id)
        .filter(Debarquement.date_debarquement.between(date_debut, date_fin))
        .group_by("annee", "mois")
        .order_by("annee", "mois")
        .all()
    )

    mois_noms = [
        "",
        "Jan",
        "Fév",
        "Mar",
        "Avr",
        "Mai",
        "Jun",
        "Jul",
        "Aoû",
        "Sep",
        "Oct",
        "Nov",
        "Déc",
    ]

    data = []
    for r in resultats:
        quantite = float(r.quantite or 0)
        valeur = float(r.valeur or 0)

        data.append(
            {
                "Année": int(r.annee),
                "Mois": mois_noms[int(r.mois)],
                "Mois N°": int(r.mois),
                "Quantité (kg)": round(quantite, 2),
                "Quantité (tonnes)": round(quantite / 1000, 3),
                "Valeur (FCFA)": round(valeur, 2),
                "Valeur (M FCFA)": round(valeur / 1000000, 3),
                "Nb débarquements": r.nb_debarquements,
                "Nb pêcheurs": r.nb_pecheurs,
                "Moyenne kg/débarquement": (
                    round(quantite / r.nb_debarquements, 2)
                    if r.nb_debarquements > 0
                    else 0
                ),
            }
        )

    headers = [
        "Année",
        "Mois",
        "Mois N°",
        "Quantité (kg)",
        "Quantité (tonnes)",
        "Valeur (FCFA)",
        "Valeur (M FCFA)",
        "Nb débarquements",
        "Nb pêcheurs",
        "Moyenne kg/débarquement",
    ]

    return data, headers


def rapport_engins_peche(db: Session, date_debut: date, date_fin: date):
    """Rapport sur l'utilisation des engins de pêche"""
    # Récupérer les bateaux avec engins et leurs captures
    resultats = (
        db.query(
            Bateau.engins_peche,
            func.count(Debarquement.id).label("nb_sorties"),
            func.sum(DetailDebarquement.quantite_kg).label("quantite"),
            func.sum(DetailDebarquement.valeur_totale).label("valeur"),
        )
        .select_from(Bateau)
        .join(Debarquement, Debarquement.bateau_id == Bateau.id)
        .join(DetailDebarquement, DetailDebarquement.debarquement_id == Debarquement.id)
        .filter(Debarquement.date_debarquement.between(date_debut, date_fin))
        .filter(Bateau.engins_peche.isnot(None))
        .filter(Bateau.engins_peche != "")
        .group_by(Bateau.engins_peche)
        .all()
    )

    # Compter par type d'engin
    engins_stats = {}
    for r in resultats:
        if r.engins_peche:
            engins_list = [e.strip() for e in r.engins_peche.split(",")]
            for engin in engins_list:
                if engin not in engins_stats:
                    engins_stats[engin] = {"sorties": 0, "quantite": 0, "valeur": 0}
                engins_stats[engin]["sorties"] += r.nb_sorties or 0
                engins_stats[engin]["quantite"] += float(r.quantite or 0)
                engins_stats[engin]["valeur"] += float(r.valeur or 0)

    data = []
    total_quantite = sum(s["quantite"] for s in engins_stats.values())

    for engin, stats in sorted(
        engins_stats.items(), key=lambda x: x[1]["quantite"], reverse=True
    ):
        pourcentage = (
            (stats["quantite"] / total_quantite * 100) if total_quantite > 0 else 0
        )

        data.append(
            {
                "Engin de pêche": engin,
                "Nb sorties": stats["sorties"],
                "Quantité (kg)": round(stats["quantite"], 2),
                "Quantité (tonnes)": round(stats["quantite"] / 1000, 3),
                "Valeur (FCFA)": round(stats["valeur"], 2),
                "% de la production": round(pourcentage, 2),
                "Moyenne kg/sortie": (
                    round(stats["quantite"] / stats["sorties"], 2)
                    if stats["sorties"] > 0
                    else 0
                ),
            }
        )

    headers = [
        "Engin de pêche",
        "Nb sorties",
        "Quantité (kg)",
        "Quantité (tonnes)",
        "Valeur (FCFA)",
        "% de la production",
        "Moyenne kg/sortie",
    ]

    return data, headers


def rapport_zones_peche(db: Session, date_debut: date, date_fin: date):
    """Rapport sur l'activité par zone de pêche"""
    resultats = (
        db.query(
            Bateau.zone_peche_habituelle,
            func.count(Debarquement.id).label("nb_sorties"),
            func.sum(DetailDebarquement.quantite_kg).label("quantite"),
            func.sum(DetailDebarquement.valeur_totale).label("valeur"),
            func.count(func.distinct(Bateau.id)).label("nb_bateaux"),
        )
        .select_from(Bateau)
        .join(Debarquement, Debarquement.bateau_id == Bateau.id)
        .join(DetailDebarquement, DetailDebarquement.debarquement_id == Debarquement.id)
        .filter(Debarquement.date_debarquement.between(date_debut, date_fin))
        .filter(Bateau.zone_peche_habituelle.isnot(None))
        .filter(Bateau.zone_peche_habituelle != "")
        .group_by(Bateau.zone_peche_habituelle)
        .order_by(func.sum(DetailDebarquement.quantite_kg).desc())
        .all()
    )

    data = []
    total_quantite = sum(float(r.quantite or 0) for r in resultats)

    for r in resultats:
        quantite = float(r.quantite or 0)
        valeur = float(r.valeur or 0)
        pourcentage = (quantite / total_quantite * 100) if total_quantite > 0 else 0

        data.append(
            {
                "Zone de pêche": r.zone_peche_habituelle or "Non spécifiée",
                "Nb sorties": r.nb_sorties,
                "Nb bateaux": r.nb_bateaux,
                "Quantité (kg)": round(quantite, 2),
                "Quantité (tonnes)": round(quantite / 1000, 3),
                "Valeur (FCFA)": round(valeur, 2),
                "% de la production": round(pourcentage, 2),
                "Moyenne kg/sortie": (
                    round(quantite / r.nb_sorties, 2) if r.nb_sorties > 0 else 0
                ),
            }
        )

    headers = [
        "Zone de pêche",
        "Nb sorties",
        "Nb bateaux",
        "Quantité (kg)",
        "Quantité (tonnes)",
        "Valeur (FCFA)",
        "% de la production",
        "Moyenne kg/sortie",
    ]

    return data, headers


def rapport_effort_peche(db: Session, date_debut: date, date_fin: date):
    """Rapport sur l'effort de pêche (CPUE - Capture Par Unité d'Effort)"""
    resultats = (
        db.query(
            Debarquement.id,
            Debarquement.date_debarquement,
            Debarquement.heure_depart,
            Debarquement.heure_arrivee,
            Pecheur.nom.label("pecheur_nom"),
            Pecheur.prenom.label("pecheur_prenom"),
            Bateau.numero_immatriculation,
            func.sum(DetailDebarquement.quantite_kg).label("quantite"),
        )
        .select_from(Debarquement)
        .join(Pecheur, Debarquement.pecheur_principal_id == Pecheur.id)
        .join(Bateau, Debarquement.bateau_id == Bateau.id)
        .join(DetailDebarquement, DetailDebarquement.debarquement_id == Debarquement.id)
        .filter(Debarquement.date_debarquement.between(date_debut, date_fin))
        .filter(Debarquement.heure_depart.isnot(None))
        .filter(Debarquement.heure_arrivee.isnot(None))
        .group_by(
            Debarquement.id,
            Debarquement.date_debarquement,
            Debarquement.heure_depart,
            Debarquement.heure_arrivee,
            Pecheur.nom,
            Pecheur.prenom,
            Bateau.numero_immatriculation,
        )
        .all()
    )

    data = []
    for r in resultats:
        # Calculer la durée en heures
        if r.heure_depart and r.heure_arrivee:
            depart_seconds = r.heure_depart.hour * 3600 + r.heure_depart.minute * 60
            arrivee_seconds = r.heure_arrivee.hour * 3600 + r.heure_arrivee.minute * 60

            # Gérer le cas où la sortie passe minuit
            if arrivee_seconds < depart_seconds:
                arrivee_seconds += 24 * 3600

            duree_heures = (arrivee_seconds - depart_seconds) / 3600
            quantite = float(r.quantite or 0)
            cpue = quantite / duree_heures if duree_heures > 0 else 0

            data.append(
                {
                    "Date": r.date_debarquement.strftime("%Y-%m-%d"),
                    "Pêcheur": f"{r.pecheur_nom} {r.pecheur_prenom}",
                    "Bateau": r.numero_immatriculation,
                    "Heure départ": r.heure_depart.strftime("%H:%M"),
                    "Heure arrivée": r.heure_arrivee.strftime("%H:%M"),
                    "Durée (h)": round(duree_heures, 2),
                    "Quantité (kg)": round(quantite, 2),
                    "CPUE (kg/h)": round(cpue, 2),
                }
            )

    headers = [
        "Date",
        "Pêcheur",
        "Bateau",
        "Heure départ",
        "Heure arrivée",
        "Durée (h)",
        "Quantité (kg)",
        "CPUE (kg/h)",
    ]

    return data, headers


def rapport_saisonnalite(db: Session, date_debut: date, date_fin: date):
    """Rapport sur la saisonnalité des captures par espèce"""
    resultats = (
        db.query(
            Espece.nom_commun_francais,
            extract("month", Debarquement.date_debarquement).label("mois"),
            func.sum(DetailDebarquement.quantite_kg).label("quantite"),
            func.count(DetailDebarquement.id).label("nb_captures"),
        )
        .select_from(Espece)
        .join(DetailDebarquement, DetailDebarquement.espece_id == Espece.id)
        .join(Debarquement, Debarquement.id == DetailDebarquement.debarquement_id)
        .filter(Debarquement.date_debarquement.between(date_debut, date_fin))
        .group_by(Espece.id, Espece.nom_commun_francais, "mois")
        .order_by(Espece.nom_commun_francais, "mois")
        .all()
    )

    mois_noms = [
        "",
        "Jan",
        "Fév",
        "Mar",
        "Avr",
        "Mai",
        "Jun",
        "Jul",
        "Aoû",
        "Sep",
        "Oct",
        "Nov",
        "Déc",
    ]

    data = []
    for r in resultats:
        quantite = float(r.quantite or 0)

        data.append(
            {
                "Espèce": r.nom_commun_francais,
                "Mois": mois_noms[int(r.mois)],
                "Mois N°": int(r.mois),
                "Quantité (kg)": round(quantite, 2),
                "Quantité (tonnes)": round(quantite / 1000, 3),
                "Nb captures": r.nb_captures,
            }
        )

    headers = [
        "Espèce",
        "Mois",
        "Mois N°",
        "Quantité (kg)",
        "Quantité (tonnes)",
        "Nb captures",
    ]

    return data, headers


def rapport_tailles_captures(
    db: Session, date_debut: date, date_fin: date, espece_id: Optional[int] = None
):
    """Rapport sur la distribution des tailles capturées"""
    query = (
        db.query(
            Espece.nom_commun_francais,
            DetailDebarquement.taille_moyenne_cm,
            func.sum(DetailDebarquement.quantite_kg).label("quantite"),
            func.count(DetailDebarquement.id).label("nb_individus"),
        )
        .select_from(Espece)
        .join(DetailDebarquement, DetailDebarquement.espece_id == Espece.id)
        .join(Debarquement, Debarquement.id == DetailDebarquement.debarquement_id)
        .filter(Debarquement.date_debarquement.between(date_debut, date_fin))
        .filter(DetailDebarquement.taille_moyenne_cm.isnot(None))
    )

    if espece_id:
        query = query.filter(Espece.id == espece_id)

    query = query.group_by(
        Espece.id, Espece.nom_commun_francais, DetailDebarquement.taille_moyenne_cm
    ).order_by(Espece.nom_commun_francais, DetailDebarquement.taille_moyenne_cm)

    resultats = query.all()

    data = []
    for r in resultats:
        quantite = float(r.quantite or 0)

        data.append(
            {
                "Espèce": r.nom_commun_francais,
                "Taille moyenne (cm)": r.taille_moyenne_cm,
                "Quantité (kg)": round(quantite, 2),
                "Nb individus": r.nb_individus,
                "Poids moyen (kg)": (
                    round(quantite / r.nb_individus, 3) if r.nb_individus > 0 else 0
                ),
            }
        )

    headers = [
        "Espèce",
        "Taille moyenne (cm)",
        "Quantité (kg)",
        "Nb individus",
        "Poids moyen (kg)",
    ]

    return data, headers


def rapport_prix_marche(db: Session, date_debut: date, date_fin: date):
    """Rapport sur l'évolution des prix au marché"""
    resultats = (
        db.query(
            extract("year", Debarquement.date_debarquement).label("annee"),
            extract("month", Debarquement.date_debarquement).label("mois"),
            Espece.nom_commun_francais,
            func.avg(DetailDebarquement.prix_unitaire_kg).label("prix_moyen"),
            func.min(DetailDebarquement.prix_unitaire_kg).label("prix_min"),
            func.max(DetailDebarquement.prix_unitaire_kg).label("prix_max"),
            func.sum(DetailDebarquement.quantite_kg).label("quantite"),
        )
        .select_from(Espece)
        .join(DetailDebarquement, DetailDebarquement.espece_id == Espece.id)
        .join(Debarquement, Debarquement.id == DetailDebarquement.debarquement_id)
        .filter(Debarquement.date_debarquement.between(date_debut, date_fin))
        .group_by("annee", "mois", Espece.id, Espece.nom_commun_francais)
        .order_by("annee", "mois", Espece.nom_commun_francais)
        .all()
    )

    mois_noms = [
        "",
        "Jan",
        "Fév",
        "Mar",
        "Avr",
        "Mai",
        "Jun",
        "Jul",
        "Aoû",
        "Sep",
        "Oct",
        "Nov",
        "Déc",
    ]

    data = []
    for r in resultats:
        data.append(
            {
                "Année": int(r.annee),
                "Mois": mois_noms[int(r.mois)],
                "Espèce": r.nom_commun_francais,
                "Prix moyen (FCFA/kg)": round(float(r.prix_moyen or 0), 2),
                "Prix min (FCFA/kg)": round(float(r.prix_min or 0), 2),
                "Prix max (FCFA/kg)": round(float(r.prix_max or 0), 2),
                "Quantité échangée (kg)": round(float(r.quantite or 0), 2),
                "Écart prix (%)": (
                    round(
                        (
                            (float(r.prix_max or 0) - float(r.prix_min or 0))
                            / float(r.prix_moyen or 1)
                            * 100
                        ),
                        2,
                    )
                    if r.prix_moyen
                    else 0
                ),
            }
        )

    headers = [
        "Année",
        "Mois",
        "Espèce",
        "Prix moyen (FCFA/kg)",
        "Prix min (FCFA/kg)",
        "Prix max (FCFA/kg)",
        "Quantité échangée (kg)",
        "Écart prix (%)",
    ]

    return data, headers


def rapport_rentabilite_sortie(db: Session, date_debut: date, date_fin: date):
    """Rapport sur la rentabilité des sorties de pêche"""
    resultats = (
        db.query(
            Debarquement.id,
            Debarquement.date_debarquement,
            Pecheur.nom.label("pecheur_nom"),
            Pecheur.prenom.label("pecheur_prenom"),
            Bateau.numero_immatriculation,
            Bateau.type_bateau,
            func.sum(DetailDebarquement.quantite_kg).label("quantite"),
            func.sum(DetailDebarquement.valeur_totale).label("valeur"),
        )
        .select_from(Debarquement)
        .join(Pecheur, Debarquement.pecheur_principal_id == Pecheur.id)
        .join(Bateau, Debarquement.bateau_id == Bateau.id)
        .join(DetailDebarquement, DetailDebarquement.debarquement_id == Debarquement.id)
        .filter(Debarquement.date_debarquement.between(date_debut, date_fin))
        .group_by(
            Debarquement.id,
            Debarquement.date_debarquement,
            Pecheur.nom,
            Pecheur.prenom,
            Bateau.numero_immatriculation,
            Bateau.type_bateau,
        )
        .order_by(func.sum(DetailDebarquement.valeur_totale).desc())
        .all()
    )

    data = []
    for r in resultats:
        quantite = float(r.quantite or 0)
        valeur = float(r.valeur or 0)

        # Catégoriser la rentabilité
        if valeur >= 500000:
            categorie = "Très rentable"
        elif valeur >= 250000:
            categorie = "Rentable"
        elif valeur >= 100000:
            categorie = "Moyen"
        else:
            categorie = "Faible"

        data.append(
            {
                "Date": r.date_debarquement.strftime("%Y-%m-%d"),
                "Pêcheur": f"{r.pecheur_nom} {r.pecheur_prenom}",
                "Bateau": r.numero_immatriculation,
                "Type bateau": r.type_bateau,
                "Quantité (kg)": round(quantite, 2),
                "Valeur (FCFA)": round(valeur, 2),
                "Valeur (k FCFA)": round(valeur / 1000, 2),
                "Catégorie": categorie,
            }
        )

    headers = [
        "Date",
        "Pêcheur",
        "Bateau",
        "Type bateau",
        "Quantité (kg)",
        "Valeur (FCFA)",
        "Valeur (k FCFA)",
        "Catégorie",
    ]

    return data, headers


def rapport_composition_captures(db: Session, date_debut: date, date_fin: date):
    """Rapport sur la diversité et composition des captures"""
    resultats = (
        db.query(
            Debarcadere.denomination,
            func.count(func.distinct(DetailDebarquement.espece_id)).label("nb_especes"),
            func.sum(DetailDebarquement.quantite_kg).label("quantite_totale"),
            func.count(Debarquement.id).label("nb_debarquements"),
        )
        .select_from(Debarcadere)
        .join(Debarquement, Debarquement.debarcadere_id == Debarcadere.id)
        .join(DetailDebarquement, DetailDebarquement.debarquement_id == Debarquement.id)
        .filter(Debarquement.date_debarquement.between(date_debut, date_fin))
        .group_by(Debarcadere.id, Debarcadere.denomination)
        .order_by(func.count(func.distinct(DetailDebarquement.espece_id)).desc())
        .all()
    )

    data = []
    for r in resultats:
        quantite = float(r.quantite_totale or 0)
        diversite_index = (
            r.nb_especes / r.nb_debarquements if r.nb_debarquements > 0 else 0
        )

        data.append(
            {
                "Débarcadère": r.denomination,
                "Nb espèces": r.nb_especes,
                "Nb débarquements": r.nb_debarquements,
                "Quantité totale (kg)": round(quantite, 2),
                "Indice diversité": round(diversite_index, 2),
                "Espèces/débarquement": (
                    round(r.nb_especes / r.nb_debarquements, 2)
                    if r.nb_debarquements > 0
                    else 0
                ),
            }
        )

    headers = [
        "Débarcadère",
        "Nb espèces",
        "Nb débarquements",
        "Quantité totale (kg)",
        "Indice diversité",
        "Espèces/débarquement",
    ]

    return data, headers


def rapport_surveillance_ressources(db: Session, date_debut: date, date_fin: date):
    """Rapport sur les indicateurs de surveillance des ressources"""
    # Espèces avec quotas
    especes_quota = (
        db.query(
            Espece.nom_commun_francais,
            Espece.quota_annuel,
            Espece.taille_minimale_capture,
            Espece.periode_reproduction,
            func.sum(DetailDebarquement.quantite_kg).label("quantite_capturee"),
            func.avg(DetailDebarquement.taille_moyenne_cm).label("taille_moyenne"),
            func.count(DetailDebarquement.id).label("nb_captures"),
        )
        .select_from(Espece)
        .join(DetailDebarquement, DetailDebarquement.espece_id == Espece.id)
        .join(Debarquement, Debarquement.id == DetailDebarquement.debarquement_id)
        .filter(Debarquement.date_debarquement.between(date_debut, date_fin))
        .group_by(
            Espece.id,
            Espece.nom_commun_francais,
            Espece.quota_annuel,
            Espece.taille_minimale_capture,
            Espece.periode_reproduction,
        )
        .all()
    )

    data = []
    for e in especes_quota:
        quantite = float(e.quantite_capturee or 0)
        quota = float(e.quota_annuel or 0)
        taille_moy = float(e.taille_moyenne or 0)
        taille_min = e.taille_minimale_capture or 0

        # Statut quota
        if quota > 0:
            pct_quota = (quantite / quota) * 100
            if pct_quota >= 100:
                statut_quota = "❌ Dépassé"
            elif pct_quota >= 90:
                statut_quota = "⚠️ Critique"
            elif pct_quota >= 75:
                statut_quota = "⚠️ Attention"
            else:
                statut_quota = "✅ Normal"
        else:
            statut_quota = "N/A"
            pct_quota = 0

        # Statut taille
        if taille_min > 0 and taille_moy > 0:
            if taille_moy >= taille_min:
                statut_taille = "✅ Conforme"
            else:
                statut_taille = "❌ Non conforme"
        else:
            statut_taille = "N/A"

        data.append(
            {
                "Espèce": e.nom_commun_francais,
                "Quota (kg)": round(quota, 2) if quota > 0 else "N/A",
                "Capturé (kg)": round(quantite, 2),
                "% Quota": round(pct_quota, 2) if quota > 0 else "N/A",
                "Statut quota": statut_quota,
                "Taille min (cm)": taille_min if taille_min > 0 else "N/A",
                "Taille moy (cm)": round(taille_moy, 2) if taille_moy > 0 else "N/A",
                "Statut taille": statut_taille,
                "Période repro": e.periode_reproduction or "N/A",
                "Nb captures": e.nb_captures,
            }
        )

    headers = [
        "Espèce",
        "Quota (kg)",
        "Capturé (kg)",
        "% Quota",
        "Statut quota",
        "Taille min (cm)",
        "Taille moy (cm)",
        "Statut taille",
        "Période repro",
        "Nb captures",
    ]

    return data, headers
